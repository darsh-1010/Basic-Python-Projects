import os
import openpyxl

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def create_event():
    clear_screen()
    print("Create a New Event\n")
    event_name = input("Enter the name of the event: ")
    event_folder = event_name.replace(" ", "_")

    if not os.path.exists(event_folder):
        os.makedirs(event_folder)

    text_file_path = os.path.join(event_folder, event_name + ".txt")
    excel_file_path = os.path.join(event_folder, event_name + ".xlsx")

    participants = []
    num_participants = int(input("Enter the number of participants: "))
    for _ in range(num_participants):
        participant = input(f"Enter participant #{len(participants)+1} name: ")
        participants.append(participant)

    participants.sort()

    sponsors = []
    num_sponsors = int(input("Enter the number of sponsors: "))
    for _ in range(num_sponsors):
        sponsor = input(f"Enter sponsor #{len(sponsors)+1} name: ")
        sponsors.append(sponsor)

    sponsors.sort()

    total_budget = float(input("Enter the total budget for the event: "))

    with open(text_file_path, "w") as text_file:
        text_file.write("Participants:\n")
        text_file.write("\n".join(participants))
        text_file.write("\n\nSponsors:\n")
        text_file.write("\n".join(sponsors))
        text_file.write(f"\n\nTotal Budget: ${total_budget}")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = event_name

    for i, participant in enumerate(participants, start=1):
        worksheet.cell(row=i, column=1, value=participant)

    for i, sponsor in enumerate(sponsors, start=1):
        worksheet.cell(row=i, column=2, value=sponsor)

    worksheet.cell(row=len(participants) + len(sponsors) + 1, column=1, value="Total Budget")
    worksheet.cell(row=len(participants) + len(sponsors) + 1, column=2, value=total_budget)

    workbook.save(excel_file_path)

    clear_screen()
    print("Event details have been successfully stored.\n")

def fetch_event_details():
    clear_screen()
    print("Fetch Event Details\n")
    event_name = input("Enter the name of the event to fetch details: ")
    event_folder = event_name.replace(" ", "_")
    text_file_path = os.path.join(event_folder, event_name + ".txt")

    if os.path.exists(text_file_path):
        with open(text_file_path, "r") as text_file:
            event_details = text_file.read()
            clear_screen()
            print("Event Details:\n")
            print(event_details)
    else:
        clear_screen()
        print("Error: No such event file found.\n")

def main():
    clear_screen()
    print("Welcome to the Event Management System!")

    while True:
        print("\nMenu:")
        print("1. Create a new event")
        print("2. Fetch details of an existing event")
        print("3. Exit")

        choice = input("Enter your choice: ")

        if choice == "1":
            create_event()
        elif choice == "2":
            fetch_event_details()
        elif choice == "3":
            clear_screen()
            print("Exiting the program. Goodbye!\n")
            break
        else:
            clear_screen()
            print("Invalid choice. Please select a valid option.\n")

if __name__ == "__main__":
    main()
